"""
Carga piloto: EN_ACS880-04_560_to_2200_kW_HW_L_A4.xlsx → Supabase (PostgreSQL).

Uso:
  copy .env.example .env   # y rellena DATABASE_URL
  pip install -r requirements.txt
  python etl/load_pilot_excel.py
"""
from __future__ import annotations

import os
import sys
import uuid
from pathlib import Path
from urllib.parse import quote_plus

import openpyxl
import psycopg2
from psycopg2.extras import Json, RealDictCursor
from dotenv import load_dotenv

from frame_parser import normalize_frame_key, parse_frame_expression

ROOT = Path(__file__).resolve().parents[1]
EXCEL_DEFAULT = Path(r"C:\Users\aleja\OneDrive\Documentos\HW\EN_ACS880-04_560_to_2200_kW_HW_L_A4.xlsx")
VERSION_CATALOGO = "EN_ACS880-04_HW_L_A4"
FAMILIA_CODIGO = "ACS880-04"
MANUAL_ORIGEN = "EN_ACS880-04_560_to_2200_kW_HW_L_A4.xlsx"


def _s(val) -> str | None:
    if val is None:
        return None
    t = str(val).strip()
    return t if t else None


def _num(val):
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _ordering_code(val) -> str | None:
    s = _s(val)
    if not s:
        return None
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _voltaje(val) -> str | None:
    s = _s(val)
    if not s:
        return None
    return s.replace(" ", "").upper()


def _validate_database_url(url: str) -> None:
    """Detecta URIs mal pegadas (error típico: postgres@password@host)."""
    if not url.startswith(("postgresql://", "postgres://")):
        raise SystemExit(
            "DATABASE_URL debe empezar por postgresql://\n"
            "Copia la URI desde Supabase → Project Settings → Database → Connection string → URI"
        )
    try:
        without_scheme = url.split("://", 1)[1]
        _, host_part = without_scheme.rsplit("@", 1)
    except ValueError:
        return
    host = host_part.split("/")[0].split(":")[0]
    if "@" in host:
        raise SystemExit(
            "\nDATABASE_URL mal formada (¿contraseña con '@' sin codificar?).\n\n"
            "Opción A — variables separadas en .env (recomendado si la contraseña tiene @):\n"
            "  DB_HOST=db.xxxxx.supabase.co\n"
            "  DB_PASSWORD=tu@contraseña\n"
            "  DB_USER=postgres\n"
            "  DB_PORT=5432\n\n"
            "Opción B — codificar @ como %40 en DATABASE_URL:\n"
            "  contraseña 'mi@clave' → 'mi%40clave'\n"
        )


def _connection_help(env_path: Path) -> str:
    example = ROOT / ".env.example"
    return (
        f"\nConfigura la base de datos en:\n  {env_path}\n\n"
        "Opción recomendada si la contraseña tiene @, # o %:\n"
        "  DB_HOST=db.TU_PROJECT_REF.supabase.co\n"
        "  DB_USER=postgres\n"
        "  DB_PASSWORD=tu contraseña tal cual (con @)\n"
        "  DB_PORT=5432\n"
        "  DB_NAME=postgres\n\n"
        f"Plantilla: {example}\n"
    )
def load_perdidas(ws, cur, variadores: dict[str, str]):
    """
    Carga hoja Perdidas:
    Columna A: U_n
    Columna B: modelo
    Columna C: Referencia interna
    Columna D: P_loss [kW]
    Columna E: Air flow [m3/h]
    Columna F: Noise [dB]
    """
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 6:
            continue

        ref = _s(row[2])  # Referencia interna

        if not ref or ref not in variadores:
            continue

        p_loss_kw = _num(row[3])
        airflow_m3h = _num(row[4])
        noise_db = _num(row[5])

        cur.execute(
            """
            UPDATE variadores
            SET
              p_loss_kw = %s,
              airflow_m3h = %s,
              noise_db = %s
            WHERE id_variador = %s
            """,
            (
                p_loss_kw,
                airflow_m3h,
                noise_db,
                variadores[ref],
            ),
        )

def get_conn():
    env_path = ROOT / ".env"
    load_dotenv(env_path)

    db_password = os.environ.get("DB_PASSWORD", "").strip()
    db_host = os.environ.get("DB_HOST", "").strip()

    if db_password and db_host:
        try:
            return psycopg2.connect(
                host=db_host,
                port=os.environ.get("DB_PORT", "5432").strip(),
                user=os.environ.get("DB_USER", "postgres").strip(),
                password=db_password,
                dbname=os.environ.get("DB_NAME", "postgres").strip(),
                sslmode=os.environ.get("DB_SSLMODE", "require").strip(),
                connect_timeout=int(os.environ.get("DB_CONNECT_TIMEOUT", "15")),
            )
        except psycopg2.OperationalError as exc:
            msg = str(exc).lower()
            if "password authentication failed" in msg:
                raise SystemExit(
                    "\nContraseña de base de datos incorrecta.\n"
                    "Supabase → Database → Reset database password → actualiza DB_PASSWORD en .env\n"
                    "(Usa la contraseña de Database, no las API keys de Settings → API)\n"
                ) from exc
            if "could not translate host name" in msg and db_host.startswith("db.") and db_host.endswith(".supabase.co"):
                raise SystemExit(
                    f"\nNo se pudo resolver el host directo:\n  {db_host}\n\n"
                    "En muchas redes Windows el host db.*.supabase.co solo tiene IPv6 y falla el DNS.\n\n"
                    "Usa Session pooler en .env (Supabase → Database → Connection string → Session pooler):\n"
                    "  DB_HOST=aws-0-TU-REGION.pooler.supabase.com\n"
                    "  DB_PORT=6543\n"
                    "  DB_USER=postgres.ylzydydcpdgshfaprpup\n"
                    "  DB_PASSWORD=(tu contraseña)\n\n"
                    "Prueba primero: python etl/test_connection.py\n"
                ) from exc
            raise

    url = os.environ.get("DATABASE_URL", "").strip()

    # URI de Supabase con [YOUR-PASSWORD] + contraseña en variable aparte
    if url and "[YOUR-PASSWORD]" in url and db_password:
        url = url.replace("[YOUR-PASSWORD]", quote_plus(db_password))

    if not url:
        raise SystemExit(_connection_help(env_path))

    _validate_database_url(url)
    return psycopg2.connect(url)


def fetch_one(cur, sql, args=()):
    cur.execute(sql, args)
    return cur.fetchone()


def ensure_lookup(cur, table: str, id_col: str, codigo_col: str, codigo: str, desc: str | None = None):
    row = fetch_one(
        cur,
        f"SELECT {id_col} FROM {table} WHERE {codigo_col} = %s",
        (codigo,),
    )
    if row:
        return row[id_col]
    new_id = str(uuid.uuid4())
    cur.execute(
        f"INSERT INTO {table} ({id_col}, {codigo_col}, descripcion) VALUES (%s, %s, %s)",
        (new_id, codigo, desc),
    )
    return new_id


def ensure_tipo_componente(cur, codigo: str) -> str:
    return ensure_lookup(cur, "tipos_componentes", "id_tipo_componente", "codigo", codigo, codigo)


def ensure_frame(cur, frame_raw: str, categoria: str = "drive") -> tuple[str, str]:
    """Retorna (id_frame, codigo_frame). Guarda frames base: D8T, R8I, R10, etc."""
    key = _s(frame_raw)
    if not key:
        raise ValueError("frame vacío")

    key = key.upper().replace(" ", "")

    row = fetch_one(cur, "SELECT id_frame FROM frames WHERE codigo_frame = %s", (key,))
    if row:
        return row["id_frame"], key

    fid = str(uuid.uuid4())
    cur.execute(
        "INSERT INTO frames (id_frame, codigo_frame, categoria) VALUES (%s, %s, %s)",
        (fid, key, categoria),
    )
    return fid, key


def ensure_componente(cur, ordering: str, nombre: str, tipo: str, descripcion: str | None = None) -> str:
    ordering = _ordering_code(ordering)
    if not ordering:
        ordering = f"NO_CODE_{uuid.uuid4().hex[:8]}"
    row = fetch_one(
        cur, "SELECT id_componente FROM componentes WHERE ordering_code = %s", (ordering,)
    )
    if row:
        return row["id_componente"]
    tid = ensure_tipo_componente(cur, tipo)
    cid = str(uuid.uuid4())
    cur.execute(
        """
        INSERT INTO componentes (id_componente, ordering_code, nombre_componente, id_tipo_componente, descripcion)
        VALUES (%s, %s, %s, %s, %s)
        """,
        (cid, ordering, nombre, tid, descripcion),
    )
    return cid


def ensure_variador(
    cur,
    *,
    referencia: str,
    modelo: str,
    voltaje: str,
    i_n,
    i_max,
    i_1,
    p_n,
    s_n,
    i_ld,
    p_ld,
    i_hd,
    p_hd,
    frame_raw: str,
    id_familia: str,
    id_import: str,
) -> str:
    hp_nominal = round(p_n * 1.34102209, 2) if p_n is not None else None

    row = fetch_one(
        cur,
        """
        SELECT id_variador 
        FROM variadores
        WHERE referencia_interna = %s 
          AND version_catalogo = %s
        """,
        (referencia, VERSION_CATALOGO),
    )

    if row:
        vid = row["id_variador"]

        cur.execute(
            """
            UPDATE variadores SET
              modelo = %s,
              voltaje_nominal = %s,
              corriente_nominal = %s,
              corriente_max = %s,
              corriente_i1_a = %s,
              potencia_nominal_kw = %s,
              potencia_heavy_duty_kw = %s,
              corriente_normal_duty_a = %s,
              corriente_light_duty_a = %s,
              corriente_heavy_duty_a = %s,
              potencia_normal_duty_kw = %s,
              potencia_light_duty_kw = %s,
              potencia_nominal_kva = %s,
              potencia_nominal_hp = %s,
              manual_origen = %s
            WHERE id_variador = %s
            """,
            (
                modelo,
                voltaje,
                i_n,
                i_max,
                i_1,
                p_n,
                p_hd,
                i_n,
                i_ld,
                i_hd,
                p_n,
                p_ld,
                s_n,
                hp_nominal,
                MANUAL_ORIGEN,
                vid,
            ),
        )

    else:
        vid = str(uuid.uuid4())

        cur.execute(
            """
            INSERT INTO variadores (
              id_variador,
              modelo,
              referencia_interna,
              id_familia,
              voltaje_nominal,
              corriente_nominal,
              corriente_max,
              corriente_i1_a,
              potencia_nominal_kw,
              potencia_heavy_duty_kw,
              corriente_normal_duty_a,
              corriente_light_duty_a,
              corriente_heavy_duty_a,
              potencia_normal_duty_kw,
              potencia_light_duty_kw,
              potencia_nominal_kva,
              potencia_nominal_hp,
              version_catalogo,
              manual_origen
            ) VALUES (
              %s, %s, %s, %s, %s, %s, %s, %s, %s,
              %s, %s, %s, %s, %s, %s, %s, %s, %s,%s
            )
            """,
            (
                vid,
                modelo,
                referencia,
                id_familia,
                voltaje,
                i_n,
                i_max,
                i_1,
                p_n,
                p_hd,
                i_n,
                i_ld,
                i_hd,
                p_n,
                p_ld,
                s_n,
                hp_nominal,
                VERSION_CATALOGO,
                MANUAL_ORIGEN,
            ),
        )

    cur.execute("DELETE FROM variador_frames WHERE id_variador = %s", (vid,))

    for tok in parse_frame_expression(frame_raw):
        fid, _ = ensure_frame(cur, tok.codigo, "drive")

        cur.execute(
            """
            INSERT INTO variador_frames (
              id_variador_frame,
              id_variador,
              id_frame,
              cantidad,
              posicion
            ) VALUES (%s, %s, %s, %s, %s)
            """,
            (
                str(uuid.uuid4()),
                vid,
                fid,
                tok.cantidad,
                tok.posicion,
            ),
        )

    cur.execute(
        """
        INSERT INTO stg_variadores_raw (
          id_raw,
          id_import,
          row_number,
          modelo_raw,
          frame_raw,
          data_json
        ) VALUES (%s, %s, %s, %s, %s, %s)
        """,
        (
            str(uuid.uuid4()),
            id_import,
            0,
            referencia,
            frame_raw,
            Json(
                {
                    "modelo": modelo,
                    "voltaje": voltaje,
                    "corriente_normal_duty_a": i_n,
                    "corriente_max": i_max,
                    "corriente_i1_a": i_1,
                    "corriente_light_duty_a": i_ld,
                    "corriente_heavy_duty_a": i_hd,
                    "potencia_normal_duty_kw": p_n,
                    "potencia_light_duty_kw": p_ld,
                    "potencia_heavy_duty_kw": p_hd,
                    "potencia_nominal_kva": s_n,
                    "potencia_nominal_hp": hp_nominal,
                }
            ),
        ),
    )

    return vid

def link_variador_componente(cur, vid: str, cid: str, qty: int, ubicacion: str, obs: str | None = None):
    cur.execute(
        """
        INSERT INTO variador_componentes (id_relacion, id_variador, id_componente, cantidad, ubicacion_funcional, observaciones)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (id_variador, id_componente, ubicacion_funcional) DO UPDATE SET
          cantidad = EXCLUDED.cantidad,
          observaciones = EXCLUDED.observaciones
        """,
        (str(uuid.uuid4()), vid, cid, qty, ubicacion or "GENERAL", obs),
    )


def _ubicacion_filtro(ubicacion: str | None, observaciones: str | None = None) -> tuple[str, str | None]:
    """Clave corta para UNIQUE + texto largo en observaciones."""
    u = (ubicacion or "GENERAL").strip()
    if len(u) <= 80:
        return u, observaciones
    extra = u
    obs = f"{observaciones} | {extra}" if observaciones else extra
    return "GENERAL", obs


def link_filtro(
    cur,
    vid: str,
    ordering: str,
    nombre: str,
    tipo: str,
    qty: int,
    ubicacion: str,
    desc: str | None,
    observaciones: str | None = None,
):
    cid = ensure_componente(cur, ordering, nombre, "FILTER_OTHER", desc)
    row = fetch_one(cur, "SELECT id_filtro FROM filtros WHERE id_componente = %s", (cid,))
    if row:
        fid = row["id_filtro"]
    else:
        fid = str(uuid.uuid4())
        cur.execute(
            "INSERT INTO filtros (id_filtro, id_componente, tipo_filtro, descripcion_tecnica) VALUES (%s,%s,%s,%s)",
            (fid, cid, tipo, desc),
        )
    ubic_key, obs = _ubicacion_filtro(ubicacion, observaciones)
    cur.execute(
        """
        INSERT INTO variador_filtros (
          id_relacion, id_variador, id_filtro, cantidad, ubicacion, tipo_instalacion, observaciones
        )
        VALUES (%s, %s, %s, %s, %s, 'catalogo', %s)
        ON CONFLICT (id_variador, id_filtro, ubicacion) DO UPDATE SET
          cantidad = EXCLUDED.cantidad,
          observaciones = EXCLUDED.observaciones
        """,
        (str(uuid.uuid4()), vid, fid, qty, ubic_key, obs),
    )


def load_rangos_operacion(ws, cur, id_familia: str, id_import: str) -> dict[str, str]:
    """referencia_interna -> id_variador"""
    mapping: dict[str, str] = {}

    # La fila 1 tiene encabezados, fila 2 unidades, fila 3 primer dato válido.
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or not row[2]:
            continue

        voltaje = _voltaje(row[0])
        modelo_suffix = _s(row[1])
        referencia = _s(row[2])
        frame_raw = _s(row[3])

        if not referencia or not frame_raw:
            continue

        modelo = modelo_suffix or referencia.split("-")[-1]

        vid = ensure_variador(
            cur,
            referencia=referencia,
            modelo=modelo,
            voltaje=voltaje or "",

            # Rangos operación
            i_n=_num(row[4]),                         # I_N
            i_max=_num(row[5]),                       # Imax
            i_1=_num(row[6]) if len(row) > 6 else None, # I_1
            p_n=_num(row[7]),                         # P_N
            s_n=_num(row[8]) if len(row) > 8 else None,   # S_N / kVA
            i_ld=_num(row[9]) if len(row) > 9 else None,  # Light-overload current
            p_ld=_num(row[10]) if len(row) > 10 else None,
            i_hd=_num(row[11]) if len(row) > 11 else None,
            p_hd=_num(row[12]) if len(row) > 12 else None,

            frame_raw=frame_raw,
            id_familia=id_familia,
            id_import=id_import,
        )

        mapping[referencia] = vid

    return mapping


def load_breakers(ws, cur, variadores: dict[str, str]):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 9:
            continue
        ref = _s(row[3])
        if not ref or ref not in variadores:
            continue
        nombre = _s(row[4]) or "Breaker"
        tipo_data = _s(row[5])
        tech = _s(row[6])
        qty = int(_num(row[7]) or 1)
        ordering = _ordering_code(row[8])
        desc = " | ".join(x for x in (tipo_data, tech) if x)
        cid = ensure_componente(cur, ordering or ref, nombre, "BREAKER", desc)
        link_variador_componente(cur, variadores[ref], cid, qty, "Main circuit", desc)


def load_fusibles(ws, cur, variadores: dict[str, str]):
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 9:
            continue
        ref = _s(row[3])
        if not ref or "ACS880" not in ref:
            continue
        if ref not in variadores:
            continue
        nombre = _s(row[4]) or "Fuse"
        fuse_type = _s(row[5])
        fuse_data = _s(row[6])
        qty = int(_num(row[7]) or 1)
        ordering = _ordering_code(row[8])
        desc = " | ".join(x for x in (fuse_type, fuse_data) if x)
        cid = ensure_componente(cur, ordering or f"{ref}-fuse", nombre, "FUSE", desc)
        link_variador_componente(cur, variadores[ref], cid, qty, "Input", desc)


def load_filtros_modo_comun(ws, cur, variadores: dict[str, str]):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 6:
            continue
        ref_field = _s(row[1])
        if not ref_field:
            continue
        if ref_field in variadores:
            ref = ref_field
        elif ref_field.startswith("ACS880") and ref_field in variadores:
            ref = ref_field
        else:
            continue
        nombre = _s(row[2]) or _s(row[3]) or "Filtro modo común"
        ordering = _ordering_code(row[4])
        qty_raw = row[5]
        try:
            qty = int(float(qty_raw)) if qty_raw is not None else 1
        except (TypeError, ValueError):
            qty = 1
        ubic = _s(row[6]) if len(row) > 6 else "GENERAL"
        desc = _s(row[3])
        link_filtro(
            cur,
            variadores[ref],
            ordering or f"FILTER-{ref}",
            nombre,
            "COMMON_MODE",
            qty,
            ubic or "GENERAL",
            desc,
        )


def load_accesorios(ws, cur):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        frame_key = normalize_frame_key(_s(row[0]))
        if not frame_key:
            continue
        row_f = fetch_one(cur, "SELECT id_frame FROM frames WHERE codigo_frame = %s", (frame_key,))
        if not row_f:
            fid, _ = ensure_frame(cur, _s(row[0]))
        else:
            fid = row_f["id_frame"]
        qty = int(_num(row[4]) or 1)
        ordering = _ordering_code(row[6])
        nombre = _s(row[7]) or "Accesorio mecánico"
        ubic = _s(row[8])
        cid = ensure_componente(cur, ordering or f"MECH-{frame_key}-{nombre[:20]}", nombre, "MECHANICAL", _s(row[9]))
        ubicacion = ubic or "GENERAL"
        cur.execute(
            """
            INSERT INTO frame_accesorios (id_relacion, id_frame, id_componente, cantidad, ubicacion, observaciones)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (id_frame, id_componente, ubicacion) DO UPDATE SET
              cantidad = EXCLUDED.cantidad,
              observaciones = EXCLUDED.observaciones
            """,
            (str(uuid.uuid4()), fid, cid, qty, ubicacion, _s(row[9])),
        )


def load_frame_specs(ws, cur):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        code = _s(row[0]).upper()
        row_f = fetch_one(cur, "SELECT id_frame FROM frames WHERE codigo_frame = %s", (code,))
        if not row_f:
            fid = str(uuid.uuid4())
            cur.execute(
                "INSERT INTO frames (id_frame, codigo_frame, categoria) VALUES (%s, %s, %s)",
                (fid, code, "module_size"),
            )
        else:
            fid = row_f["id_frame"]
        cur.execute(
            """
            INSERT INTO frame_specs (
              id_spec, id_frame, alto_mm, ancho_mm, profundidad_mm, peso_kg,
              grado_proteccion, version_catalogo
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id_frame, version_catalogo) DO UPDATE SET
              alto_mm = EXCLUDED.alto_mm,
              ancho_mm = EXCLUDED.ancho_mm,
              profundidad_mm = EXCLUDED.profundidad_mm,
              peso_kg = EXCLUDED.peso_kg,
              grado_proteccion = EXCLUDED.grado_proteccion
            """,
            (
                str(uuid.uuid4()),
                fid,
                _num(row[1]),
                _num(row[2]),
                _num(row[3]),
                _num(row[4]),
                _s(row[5]),
                VERSION_CATALOGO,
            ),
        )


def load_control_sheet(ws, cur, familia_modulo: str):
    """BCU / UCU: catálogo de módulos y terminales (sin enlace a variador en este manual)."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return
    header = None
    for i, row in enumerate(rows):
        if row and _s(row[0]) == "Modulo_ID":
            header = i
            break
    if header is None:
        return
    for row in rows[header + 1 :]:
        if not row or not row[0]:
            continue
        codigo = _s(row[0])
        nombre = _s(row[1])
        if not codigo:
            continue
        row_m = fetch_one(cur, "SELECT id_control_module FROM control_modules WHERE codigo_modulo = %s", (codigo,))
        if row_m:
            mid = row_m["id_control_module"]
        else:
            mid = str(uuid.uuid4())
            cur.execute(
                """
                INSERT INTO control_modules (id_control_module, codigo_modulo, familia, tipo, descripcion)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (mid, codigo, familia_modulo, "control", nombre),
            )
        t_no = _s(row[2])
        t_lbl = _s(row[3])
        if not t_no and not t_lbl:
            continue
        cur.execute(
            """
            INSERT INTO terminales_control (
              id_terminal, id_control_module, terminal_no, terminal_etiqueta,
              descripcion_funcion, voltage_range, corriente_max
            ) VALUES (%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (id_control_module, terminal_no, terminal_etiqueta) DO UPDATE SET
              descripcion_funcion = EXCLUDED.descripcion_funcion
            """,
            (
                str(uuid.uuid4()),
                mid,
                t_no,
                t_lbl,
                _s(row[4]),
                _s(row[5]),
                _s(row[6]),
            ),
        )


def schema_exists(cur) -> bool:
    cur.execute(
        "SELECT 1 FROM information_schema.tables WHERE table_schema = 'public' AND table_name = 'variadores'"
    )
    return cur.fetchone() is not None


def apply_schema_patches(conn):
    """Parches idempotentes (columnas nuevas / tipos TEXT)."""
    patch = ROOT / "supabase" / "migrations" / "003_widen_text_columns.sql"
    if not patch.exists():
        return
    sql = patch.read_text(encoding="utf-8")
    with conn.cursor() as cur:
        cur.execute(sql)
    conn.commit()
    print("Parche de esquema aplicado (003_widen_text_columns).")


def apply_migrations(conn):
    mig_dir = ROOT / "supabase" / "migrations"
    with conn.cursor() as cur:
        if schema_exists(cur):
            print("Esquema base ya existe; aplicando solo parches...")
            apply_schema_patches(conn)
            return
    for name in sorted(mig_dir.glob("*.sql")):
        sql = name.read_text(encoding="utf-8")
        with conn.cursor() as cur:
            cur.execute(sql)
    conn.commit()
    print(f"Migraciones aplicadas desde {mig_dir}")


def main():
    excel_path = Path(os.environ.get("EXCEL_PATH", EXCEL_DEFAULT))
    if not excel_path.exists():
        raise SystemExit(f"No existe el Excel: {excel_path}")

    conn = get_conn()
    try:
        if os.environ.get("APPLY_MIGRATIONS", "1") == "1":
            apply_migrations(conn)

        print(f"Cargando Excel: {excel_path.name} ...", flush=True)

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            id_familia = ensure_lookup(
                cur, "familias", "id_familia", "codigo", FAMILIA_CODIGO, "ACS880-04 cabinet drives"
            )
            id_import = str(uuid.uuid4())
            cur.execute(
                """
                INSERT INTO etl_imports (id_import, archivo_origen, version_catalogo, estado, observaciones)
                VALUES (%s, %s, %s, 'en_proceso', %s)
                """,
                (id_import, excel_path.name, VERSION_CATALOGO, "Carga piloto ETL"),
            )

            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            variadores = load_rangos_operacion(wb["Rangos_operacion"], cur, id_familia, id_import)
            print(f"  Variadores: {len(variadores)}", flush=True)

            if "breakers" in wb.sheetnames:
                load_breakers(wb["breakers"], cur, variadores)
            if "Fusibles" in wb.sheetnames:
                load_fusibles(wb["Fusibles"], cur, variadores)
            if "Filtros_modo_comun" in wb.sheetnames:
                print("  Filtros modo común...", flush=True)
                load_filtros_modo_comun(wb["Filtros_modo_comun"], cur, variadores)
            if "Perdidas" in wb.sheetnames:
                print("  Perdidas / Airflow / Noise...", flush=True)
                load_perdidas(wb["Perdidas"], cur, variadores)
            if "accesoriosMecanicos" in wb.sheetnames:
                load_accesorios(wb["accesoriosMecanicos"], cur)
            if "tamañosFrames" in wb.sheetnames:
                load_frame_specs(wb["tamañosFrames"], cur)
            elif "tamanosFrames" in wb.sheetnames:
                load_frame_specs(wb["tamanosFrames"], cur)
            for sheet, fam in (("BCU", "BCU"), ("UCU", "UCU")):
                if sheet in wb.sheetnames:
                    load_control_sheet(wb[sheet], cur, fam)

            wb.close()

            total = len(variadores)
            cur.execute(
                "UPDATE etl_imports SET estado = 'completado', registros_importados = %s WHERE id_import = %s",
                (total, id_import),
            )
            conn.commit()
            print(f"OK: {total} variadores cargados desde {excel_path.name}")
            print("Prueba en SQL Editor:")
            print("  SELECT * FROM buscar_variador('1140A-3', '400');")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent))
    main()
